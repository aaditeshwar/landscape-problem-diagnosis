#!/usr/bin/env python3
"""Compare croppingDrought_kharif Excel columns with Mongo drought_kharif for sample MWS."""

from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "runtime"))

from openpyxl import load_workbook
from pymongo import MongoClient

from config import DB_NAME, MONGO_URI

SAMPLES = [
    ("Rajasthan__Bundi__Hindoli_data.xlsx", "12_79359"),
    ("Maharashtra__Yavatmal__Darwha_data.xlsx", "4_100672"),
    ("Karnataka__Bagalkot__Hungund_data.xlsx", "18_31133"),
    ("Madhya Pradesh__Sehore__Sehore_data.xlsx", "12_112746"),
    ("Odisha__Koraput__Koraput_data.xlsx", "22_10810"),
    ("Andhra Pradesh__Ananthapur__Amadagur_data.xlsx", "3_11221"),
    ("Goa__South Goa__Sanguem_data.xlsx", "25_10003"),
    ("Jharkhand__Giridih__Gande_data.xlsx", "12_306338"),
    ("Karnataka__Kolar__Mulbagal_data.xlsx", "2_29022"),
    ("Maharashtra__Amaravati__Warud_data.xlsx", "4_100348"),
    ("Meghalaya__East Khasi Hills__East Khasi Hills_data.xlsx", "10_1"),
]


def sheet_rows(wb, name: str) -> list[dict]:
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def compare(uid: str, excel_row: dict, mongo_row: dict[str, dict]) -> list[str]:
    issues: list[str] = []
    for y in sorted(mongo_row.keys()):
        m = mongo_row[y]
        ex = {
            "no_drought_weeks": excel_row.get(f"No_Drought_in_weeks_{y}"),
            "mild_weeks": excel_row.get(f"Mild_in_weeks_{y}"),
            "moderate_weeks": excel_row.get(f"Moderate_in_weeks_{y}"),
            "severe_weeks": excel_row.get(f"Severe_in_weeks_{y}"),
            "total_weeks": excel_row.get(f"total_weeks_{y}"),
        }
        for key in ("no_drought_weeks", "mild_weeks", "moderate_weeks", "severe_weeks"):
            ev = ex[key]
            mv = m.get(key)
            if ev is None and mv is None:
                continue
            if ev is None or mv is None or float(ev) != float(mv):
                issues.append(f"{y} {key}: excel={ev} mongo={mv}")
        total = sum(float(m.get(k) or 0) for k in ("no_drought_weeks", "mild_weeks", "moderate_weeks", "severe_weeks"))
        tw = m.get("total_weeks")
        if tw is not None and abs(total - float(tw)) > 0.01:
            issues.append(f"{y} sum={total} vs total_weeks={tw}")
    return issues


def main() -> int:
    client = MongoClient(MONGO_URI)
    col = client[DB_NAME]["mws_data"]
    report = []
    for filename, uid_hint in SAMPLES:
        path = ROOT / "data" / "raw_excel" / filename
        if not path.is_file():
            report.append({"file": filename, "error": "excel missing"})
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        rows = sheet_rows(wb, "croppingDrought_kharif")
        wb.close()
        uid = uid_hint or str(rows[0].get("UID"))
        doc = col.find_one({"uid": uid}, {"drought_kharif": 1, "tehsil": 1, "district": 1, "state": 1})
        if not doc:
            report.append({"uid": uid, "file": filename, "error": "mongo missing"})
            continue
        excel_row = next((r for r in rows if str(r.get("UID")) == uid), None)
        if not excel_row:
            report.append({"uid": uid, "file": filename, "error": "uid not in excel sheet"})
            continue
        dk = doc.get("drought_kharif") or {}
        latest = sorted(dk.keys())[-1] if dk else None
        latest_row = dk.get(latest, {}) if latest else {}
        report.append(
            {
                "uid": uid,
                "location": f"{doc.get('tehsil')}, {doc.get('district')}, {doc.get('state')}",
                "file": filename,
                "latest_year": latest,
                "latest_weeks": {
                    k: latest_row.get(k)
                    for k in (
                        "no_drought_weeks",
                        "mild_weeks",
                        "moderate_weeks",
                        "severe_weeks",
                        "total_weeks",
                    )
                },
                "mismatch_count": len(issues := compare(uid, excel_row, dk)),
                "mismatches": issues[:6],
            }
        )
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
