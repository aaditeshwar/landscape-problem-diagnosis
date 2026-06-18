"""
ingest_excel.py
===============
Ingests a tehsil-level Excel data file into MongoDB (AtlasLocalDev),
fetches MWS and village geometries from the CoRE Stack API, and builds
a dissolved tehsil boundary polygon.

Usage:
    python ingest_excel.py \
        --excel /path/to/Darwha_data.xlsx \
        --state Maharashtra \
        --district Yavatmal \
        --tehsil Darwha

Environment variables (or .env file):
    CORE_STACK_API_KEY   CoRE Stack API key (X-API-Key header)
    MONGO_URI            MongoDB URI (default: mongodb://localhost:27017)
"""

import argparse
import ast
import json
import logging
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv
from pymongo import MongoClient, UpdateOne
from shapely.geometry import mapping, shape
from shapely.ops import unary_union
from shapely.validation import make_valid

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

_RUNTIME_DIR = Path(__file__).resolve().parents[1] / "runtime"
if str(_RUNTIME_DIR) not in sys.path:
    sys.path.insert(0, str(_RUNTIME_DIR))
from services.aer_lookup import attach_aer_to_mws, validate_aer_geojson  # noqa: E402
from services.aquifer_classification import (  # noqa: E402
    LITHOLOGY_COLUMNS,
    build_aquifer_payload,
    infer_acwadam_class,
)

# ── Configuration ─────────────────────────────────────────────────────────────
CORE_STACK_BASE = "https://geoserver.core-stack.org/api/v1"
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017")
DB_NAME = "diagnosis_db"

NREGA_MWS_CATEGORIES = [
    "Soil and water conservation",
    "Land restoration",
    "Plantations",
    "Irrigation on farms",
    "Other farm works",
    "Off-farm livelihood assets",
    "Community assets",
]

NREGA_VILLAGE_CATEGORIES = [
    "Community assets",
    "Irrigation on farms",
    "Land restoration",
    "Off-farm livelihood assets",
    "Other farm works",
    "Plantations",
    "Soil and water conservation",
]

HYDROLOGICAL_YEARS = [
    "2017-2018", "2018-2019", "2019-2020", "2020-2021",
    "2021-2022", "2022-2023", "2023-2024", "2024-2025",
]

LULC_YEARS = [str(y) for y in range(2017, 2025)]  # 2017..2024

DROUGHT_YEARS = [str(y) for y in range(2017, 2025)]  # 2017..2024

CROPPING_YEARS = [
    "2017-2018", "2018-2019", "2019-2020", "2020-2021",
    "2021-2022", "2022-2023", "2023-2024", "2024-2025",
]

SWB_YEARS = [
    "2017-2018", "2018-2019", "2019-2020", "2020-2021",
    "2021-2022", "2022-2023", "2023-2024", "2024-2025",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def yr(year_str: str) -> str:
    """Convert '2017-2018' to agricultural year key '2017' (BSON requires string keys)."""
    return year_str.split("-")[0]


def safe(val):
    """Return None for blank/missing values."""
    if val is None or val == "" or val != val:  # NaN check
        return None
    return val


def bson_safe(obj):
    """Recursively convert dict keys to strings for MongoDB BSON compatibility."""
    if isinstance(obj, dict):
        return {str(k): bson_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [bson_safe(v) for v in obj]
    return obj


def parse_json_field(val):
    """Parse a JSON string field, return dict or {}."""
    if not val:
        return {}
    try:
        return json.loads(val)
    except Exception:
        try:
            return ast.literal_eval(val)
        except Exception:
            return {}


def sheet_to_dict(wb, sheet_name, *, optional=False):
    """Load a worksheet and return list of row dicts keyed by header."""
    if sheet_name not in wb.sheetnames:
        if optional:
            log.info(f"  Optional sheet '{sheet_name}' absent — skipping")
            return []
        raise KeyError(f"Worksheet {sheet_name} does not exist.")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


# ── Sheet parsers ─────────────────────────────────────────────────────────────

def parse_mws(rows):
    """mws sheet → {uid: doc}"""
    docs = {}
    for r in rows:
        uid = r["UID"]
        if not uid:
            continue
        docs[uid] = {
            "uid": uid,
            "area_ha": safe(r["area_in_ha"]),
            "watershed_code": safe(r["watershed_code"]),
            "basin_code": safe(r["basin_code"]),
            "sub_basin_code": safe(r["sub_basin_code"]),
        }
    return docs


def parse_stream_order(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        docs[uid]["stream_order_area_percent"] = {
            str(i): safe(r.get(f"order_{i}_area_percent")) for i in range(1, 12)
        }


def parse_connectivity(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        upstream_raw = r.get("upstream_mws", "[]") or "[]"
        try:
            upstream = ast.literal_eval(str(upstream_raw))
        except Exception:
            upstream = []
        docs[uid]["downstream_uid"] = safe(r.get("downstream_mws"))
        docs[uid]["upstream_uids"] = upstream
        docs[uid]["flow_direction"] = safe(r.get("direction"))


def parse_terrain(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        docs[uid]["terrain"] = {
            "cluster_id": safe(r.get("terrain_cluster_id")),
            "description": safe(r.get("terrain_description")),
            "hill_slope_percent": safe(r.get("hill_slope_area_percent")),
            "plain_percent": safe(r.get("plain_area_percent")),
            "ridge_percent": safe(r.get("ridge_area_percent")),
            "slopy_percent": safe(r.get("slopy_area_percent")),
            "valley_percent": safe(r.get("valley_area_percent")),
        }


def parse_terrain_lulc(rows, docs, terrain_type):
    """terrain_type: 'slope' or 'plain'"""
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        key = f"terrain_lulc_{terrain_type}"
        docs[uid][key] = {
            "cluster_id": safe(r.get("terrain_cluster_id")),
            "cluster_name": safe(r.get("cluster_name")),
            "barren_percent": safe(r.get("barren_area_percent")),
            "forest_percent": safe(r.get("forests_area_percent")),
            "shrub_scrub_percent": safe(r.get("shrub_scrubs_area_percent")),
            "single_kharif_percent": safe(r.get("single_kharif_area_percent")),
            "single_non_kharif_percent": safe(r.get("single_non_kharif_area_percent")),
            "double_crop_percent": safe(r.get("double_cropping_area_percent")),
            "triple_crop_percent": safe(r.get("triple_cropping_area_percent")),
        }


def parse_dem(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        docs[uid]["elevation"] = {
            "min_m": safe(r.get("min_elevation")),
            "max_m": safe(r.get("max_elevation")),
            "mean_m": safe(r.get("mean_elevation")),
        }


def parse_drainage_density(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        raw = safe(r.get("drainage_density"))
        docs[uid]["drainage_density"] = {
            "raw": raw,
            "corrected_km_per_km2": round(raw / 100, 4) if raw is not None else None,
        }


def parse_aquifer(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        raw_class = safe(r.get("aquifer_class")) or ""
        lithology = {
            lit: safe(r.get(f"principle_aq_{lit}_percent"))
            for lit in LITHOLOGY_COLUMNS
        }
        docs[uid]["aquifer"] = build_aquifer_payload(raw_class, lithology)


def refine_aquifer_acwadam(db, uids: list[str]) -> dict[str, int]:
    """Recompute aquifer.acwadam_class using lithology + persisted nbss_lup_aer_code."""
    stats = {"requested": len(uids), "updated": 0, "missing": 0}
    if not uids:
        return stats

    ops: list[UpdateOne] = []
    for doc in db.mws_data.find(
        {"uid": {"$in": uids}},
        {"uid": 1, "aquifer": 1, "nbss_lup_aer_code": 1},
    ):
        uid = doc.get("uid")
        aquifer = doc.get("aquifer") or {}
        lithology = aquifer.get("lithology_percent") or {}
        if not lithology:
            stats["missing"] += 1
            continue
        inferred = infer_acwadam_class(lithology, doc.get("nbss_lup_aer_code"))
        ops.append(
            UpdateOne(
                {"uid": uid},
                {
                    "$set": {
                        "aquifer.dominant_lithology": inferred["dominant_lithology"],
                        "aquifer.acwadam_class": inferred["acwadam_class"],
                        "aquifer.acwadam_source": inferred["acwadam_source"],
                    }
                },
            )
        )
        stats["updated"] += 1

    if ops:
        db.mws_data.bulk_write(ops)
    return stats


def parse_soge(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        docs[uid]["soge"] = {
            "dev_percent": safe(r.get("soge_dev_percent")),
            "class_code": safe(r.get("class_code")),
            "class_name": safe(r.get("class_name")),
        }


def parse_restoration(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        docs[uid]["restoration_ha"] = {
            "widescale": safe(r.get("wide_scale_restoration_area_in_ha")),
            "protection": safe(r.get("protection_area_in_ha")),
            "mosaic": safe(r.get("mosaic_restoration_area_in_ha")),
            "excluded": safe(r.get("excluded_areas_in_ha")),
        }


def parse_change_detection(rows, docs, sheet_key, col_map):
    """Generic change detection parser. col_map: {output_key: source_col}"""
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        if "change_detection" not in docs[uid]:
            docs[uid]["change_detection"] = {}
        docs[uid]["change_detection"][sheet_key] = {
            k: safe(r.get(v)) for k, v in col_map.items()
        }


def parse_river(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        docs[uid]["river_name"] = safe(r.get("river_name"))


def parse_canal(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        docs[uid]["canal"] = {
            "project_name": safe(r.get("project_name")),
            "canal_code": safe(r.get("canal_code")),
            "canal_name": safe(r.get("canal_name")),
        }


def parse_hydrological_annual(rows, docs):
    """
    Compute delta_g = P - ET - Runoff for each year.
    Ignore raw G_in_mm and DeltaG_in_mm columns (buggy).
    Well depth and cumulative G are not stored (user-supplied / runtime-derived only).
    """
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        hydro = {}
        for ys in HYDROLOGICAL_YEARS:
            y = yr(ys)
            et = safe(r.get(f"ET_in_mm_{ys}")) or 0.0
            runoff = safe(r.get(f"RunOff_in_mm_{ys}")) or 0.0
            precip = safe(r.get(f"Precipitation_in_mm_{ys}")) or 0.0
            delta_g = round(precip - et - runoff, 3)
            hydro[y] = {
                "precipitation_mm": precip,
                "et_mm": et,
                "runoff_mm": runoff,
                "delta_g_mm": delta_g,
            }
        docs[uid]["hydrological_annual"] = hydro


def parse_hydrological_seasonal(rows, docs):
    """
    Compute seasonal delta_g = P - ET - Runoff per season.
    Ignore raw 'delta g_' and 'g_' columns (buggy).
    """
    seasons = ["kharif", "rabi", "zaid"]
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        seasonal = {}
        for ys in HYDROLOGICAL_YEARS:
            y = yr(ys)
            seasonal[y] = {}
            for s in seasons:
                p = safe(r.get(f"precipitation_{s}_in_mm_{ys}")) or 0.0
                et = safe(r.get(f"et_{s}_in_mm_{ys}")) or 0.0
                ro = safe(r.get(f"runoff_{s}_in_mm_{ys}")) or 0.0
                dg = round(p - et - ro, 3)
                seasonal[y][s] = {
                    "precipitation_mm": p,
                    "et_mm": et,
                    "runoff_mm": ro,
                    "delta_g_mm": dg,
                }
        docs[uid]["hydrological_seasonal"] = seasonal


def parse_swb_annual(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        swb = {}
        for ys in SWB_YEARS:
            y = yr(ys)
            swb[y] = {
                "total_ha": safe(r.get(f"total_area_in_ha_{ys}")),
                "kharif_ha": safe(r.get(f"kharif_area_in_ha_{ys}")),
                "rabi_ha": safe(r.get(f"rabi_area_in_ha_{ys}")),
                "zaid_ha": safe(r.get(f"zaid_area_in_ha_{ys}")),
            }
        docs[uid]["swb_annual"] = swb
        docs[uid]["swb_count_approx"] = safe(r.get("total_swb_area_in_ha"))


def parse_lulc_vector(rows, docs):
    lulc_fields = [
        "barrenland", "built_up_area", "cropland", "double_crop",
        "triple_crop", "tree_forest", "shrub_scrub",
        "single_kharif", "single_non_kharif",
        "k_water", "kr_water", "krz_water",
    ]
    # cropland_in_ha is stored but not used for diagnosis; lulc_cropland_ha sums the four crop classes.
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        lulc = {}
        for y_str in LULC_YEARS:
            lulc[y_str] = {
                f.replace("_area", ""): safe(r.get(f"{f}_in_ha_{y_str}"))
                for f in lulc_fields
            }
        docs[uid]["lulc_ha"] = lulc


def parse_cropping_intensity(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        ci = {}
        for ys in CROPPING_YEARS:
            y = yr(ys)
            ci[y] = {
                "cropping_intensity": safe(r.get(f"cropping_intensity_unit_less_{ys}")),
                "single_crop_ha": safe(r.get(f"single_cropped_area_in_ha_{ys}")),
                "single_kharif_ha": safe(r.get(f"single_kharif_cropped_area_in_ha_{ys}")),
                "single_non_kharif_ha": safe(r.get(f"single_non_kharif_cropped_area_in_ha_{ys}")),
                "double_crop_ha": safe(r.get(f"doubly_cropped_area_in_ha_{ys}")),
                "triple_crop_ha": safe(r.get(f"triply_cropped_area_in_ha_{ys}")),
            }
        docs[uid]["cropping_intensity"] = ci


def parse_drought_kharif(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        drought = {}
        for y_str in DROUGHT_YEARS:
            sqkm = safe(r.get(f"kharif_cropped_sqkm_{y_str}"))
            drought[y_str] = {
                "no_drought_weeks": safe(r.get(f"No_Drought_in_weeks_{y_str}")),
                "mild_weeks": safe(r.get(f"Mild_in_weeks_{y_str}")),
                "moderate_weeks": safe(r.get(f"Moderate_in_weeks_{y_str}")),
                "severe_weeks": safe(r.get(f"Severe_in_weeks_{y_str}")),
                "dry_spell_weeks": safe(r.get(f"drysp_unit_4_weeks_{y_str}")),
                "kharif_cropped_sqkm": sqkm,
                "kharif_cropped_ha": round(sqkm * 100, 4) if sqkm is not None else None,
                "kharif_cropped_percent": safe(r.get(f"kharif_cropped_area_percent_{y_str}")),
                "monsoon_onset": safe(r.get(f"monsoon_onset_{y_str}")),
                "total_weeks": safe(r.get(f"total_weeks_{y_str}")),
            }
        docs[uid]["drought_kharif"] = drought


def parse_drought_causality(rows, docs):
    for r in rows:
        uid = r["UID"]
        if uid not in docs:
            continue
        causality = {}
        for y_str in DROUGHT_YEARS:
            sev = parse_json_field(r.get(f"severe_moderate_drought_causality_{y_str}"))
            mild = parse_json_field(r.get(f"mild_drought_causality_{y_str}"))
            causality[y_str] = {"severe_moderate": sev, "mild": mild}
        docs[uid]["drought_causality"] = causality


def finalize_mws_docs(docs: dict) -> None:
    """Apply registry normalizers to parsed MWS documents before Mongo upsert."""
    from services.variable_registry import normalize_drought_causality

    for doc in docs.values():
        if "drought_causality" in doc:
            doc["drought_causality"] = normalize_drought_causality(doc.get("drought_causality"))


def parse_nrega_mws(rows, docs):
    nrega_years = list(range(2005, 2026))
    for r in rows:
        uid = r.get("mws_id")
        if uid not in docs:
            continue
        nrega = {}
        for y in nrega_years:
            nrega[str(y)] = {
                cat.lower().replace(" ", "_").replace("-", "_"): safe(r.get(f"{cat}_count_{y}"))
                for cat in NREGA_MWS_CATEGORIES
            }
        docs[uid]["nrega_mws"] = nrega


def parse_mws_intersect_villages(rows, docs):
    for r in rows:
        uid = r.get("MWS UID")
        if uid not in docs:
            continue
        raw_ids = r.get("Village IDs", "[]") or "[]"
        raw_details = r.get("Village Details", "{}") or "{}"
        try:
            village_ids = ast.literal_eval(str(raw_ids))
        except Exception:
            village_ids = []
        try:
            village_details = ast.literal_eval(str(raw_details))
        except Exception:
            village_details = {}
        docs[uid]["intersect_villages"] = {
            "village_ids": village_ids,
            "details": village_details,
        }


def parse_mws_intersect_swb(rows, docs):
    # Multiple rows per UID (one per SWB)
    swb_map = {}
    for r in rows:
        uid = r["UID"]
        if not uid:
            continue
        swb_map.setdefault(uid, []).append({
            "swb_uid": r.get("SWB_UID"),
            "name": r.get("Waterbodies_name"),
            "lat": r.get("Latitude"),
            "lon": r.get("Longitude"),
        })
    for uid, swbs in swb_map.items():
        if uid in docs:
            docs[uid]["swb_intersect"] = swbs
            docs[uid]["swb_count"] = len(swbs)


def parse_agroecological(rows, docs):
    # Multiple rows per UID (one per org)
    org_map = {}
    for r in rows:
        uid = r["UID"]
        if not uid:
            continue
        domains_raw = r.get("domains", "[]") or "[]"
        try:
            domains = json.loads(domains_raw)
        except Exception:
            domains = []
        org_map.setdefault(uid, []).append({
            "name": r.get("organization_name"),
            "type": r.get("organization_type"),
            "contact": r.get("contact_person"),
            "email": r.get("email"),
            "domains": domains,
        })
    for uid, orgs in org_map.items():
        if uid in docs:
            docs[uid]["organisations"] = orgs
            all_domains = list({d for org in orgs for d in org["domains"]})
            docs[uid]["organisation_domains"] = all_domains


# ── Village parsers ───────────────────────────────────────────────────────────

def parse_social_economic(wb):
    rows = sheet_to_dict(wb, "social_economic_indicator")
    village_docs = {}
    for r in rows:
        vid = r.get("village_id")
        if vid == 0 or vid is None:
            continue  # Exclude placeholder BANAYAT (JUNE)
        village_docs[vid] = {
            "village_id": vid,
            "village_name": r.get("village_name"),
            "state_census_id": r.get("state_census_ID"),
            "dist_census_id": r.get("dist_census_ID"),
            "block_census_id": r.get("block_census_ID"),
            "population": safe(r.get("total_population_count")),
            "sc_count": safe(r.get("total_SC_population_count")),
            "st_count": safe(r.get("total_ST_population_count")),
            "sc_percent": safe(r.get("SC_percent")),
            "st_percent": safe(r.get("ST_percent")),
            "literacy_rate_percent": safe(r.get("literacy_rate_percent")),
        }
    return village_docs


def parse_facilities_proximity(wb, village_docs):
    rows = sheet_to_dict(wb, "facilities_proximity", optional=True)
    if not rows:
        return
    for r in rows:
        vid = r.get("censuscode2011")
        if vid == 0 or vid is None:
            continue
        if vid not in village_docs:
            # Create minimal village doc if not in social sheet
            village_docs[vid] = {"village_id": vid, "village_name": r.get("censusname")}
        village_docs[vid]["facility_distances_km"] = {
            "apmc": safe(r.get("apmc_distance")),
            "bank_branch": safe(r.get("bank_branch_distance")),
            "bank_atm": safe(r.get("bank_atm_distance")),
            "bank_mitra": safe(r.get("bank_mitra_distance")),
            "phc": safe(r.get("health_phc_distance")),
            "chc": safe(r.get("health_chc_distance")),
            "sub_centre": safe(r.get("health_sub_cen_distance")),
            "district_hospital": safe(r.get("health_dis_h_distance")),
            "school_primary": safe(r.get("school_primary_distance")),
            "school_upper_primary": safe(r.get("school_upper_primary_distance")),
            "school_secondary": safe(r.get("school_secondary_distance")),
            "school_higher_secondary": safe(r.get("school_higher_secondary_distance")),
            "college": safe(r.get("college_distance")),
            "university": safe(r.get("universities_distance")),
            "csc": safe(r.get("csc_distance")),
            "pds": safe(r.get("pds_distance")),
            "agri_processing": safe(r.get("agri_industry_agri_processing_distance")),
            "agri_support": safe(r.get("agri_industry_agri_support_infrastructure_distance")),
            "cooperative": safe(r.get("agri_industry_co_operatives_societies_distance")),
            "dairy": safe(r.get("agri_industry_dairy_animal_husbandry_distance")),
            "markets_trading": safe(r.get("agri_industry_markets_trading_distance")),
            "storage_warehousing": safe(r.get("agri_industry_storage_warehousing_distance")),
        }
        village_docs[vid]["core_admin_uid"] = safe(r.get("core_admin_uid"))


def parse_nrega_village(wb, village_docs):
    rows = sheet_to_dict(wb, "nrega_assets_village", optional=True)
    if not rows:
        return
    nrega_years = list(range(2005, 2026))
    for r in rows:
        vid = r.get("vill_id")
        if vid == 0 or vid is None:
            continue
        if vid not in village_docs:
            village_docs[vid] = {"village_id": vid, "village_name": r.get("vill_name")}
        nrega = {}
        for y in nrega_years:
            nrega[str(y)] = {
                cat.lower().replace(" ", "_").replace("-", "_"): safe(r.get(f"{cat}_count_{y}"))
                for cat in NREGA_VILLAGE_CATEGORIES
            }
        village_docs[vid]["nrega_village"] = nrega


# ── CoRE Stack geometry fetch ─────────────────────────────────────────────────

def _log_geometry_ingest_stats(
    layer: str,
    features: list,
    *,
    id_keys: tuple[str, ...],
    stored: int,
    valid_shapes: int | None = None,
) -> None:
    """Log counts of features, IDs, and geometries retrieved from a GeoJSON layer."""
    total = len(features)
    with_id = 0
    with_geom = 0
    with_both = 0
    id_only = 0
    geom_only = 0
    sample_prop_keys: list[str] | None = None
    missing_id_samples: list[dict] = []

    for feat in features:
        props = feat.get("properties") or {}
        if sample_prop_keys is None and props:
            sample_prop_keys = sorted(props.keys())

        vid = None
        for key in id_keys:
            val = props.get(key)
            if val is not None and val != "":
                vid = val
                break

        geom = feat.get("geometry")
        has_id = vid is not None and vid != ""
        has_geom = geom is not None

        if has_id:
            with_id += 1
        if has_geom:
            with_geom += 1
        if has_id and has_geom:
            with_both += 1
        elif has_id:
            id_only += 1
        elif has_geom:
            geom_only += 1
            if len(missing_id_samples) < 3:
                missing_id_samples.append({
                    "property_keys": sorted(props.keys()),
                    "properties": {k: props.get(k) for k in id_keys},
                })

    log.info(f"  {layer}: {total} features in FeatureCollection")
    log.info(
        f"  {layer}: {with_id} with non-null id "
        f"({', '.join(id_keys)}), {with_geom} with geometry, {with_both} with both"
    )
    if id_only:
        log.warning(f"  {layer}: {id_only} features have id but no geometry")
    if geom_only:
        log.warning(f"  {layer}: {geom_only} features have geometry but no id")
    if sample_prop_keys:
        log.info(f"  {layer}: sample property keys from first feature: {sample_prop_keys}")
    if missing_id_samples:
        log.warning(f"  {layer}: examples of geometry-only features (no id matched):")
        for sample in missing_id_samples:
            log.warning(f"    keys={sample['property_keys']} id_fields={sample['properties']}")
    log.info(f"  {layer}: stored {stored} boundary documents in MongoDB")
    if valid_shapes is not None:
        log.info(f"  {layer}: {valid_shapes} valid shapes used for tehsil dissolve")


def geojson_for_mongo(geom):
    """Return GeoJSON dict safe for MongoDB 2dsphere index, or None if unusable."""
    if geom is None or geom.is_empty:
        return None
    fixed = make_valid(geom)
    if fixed.is_empty:
        return None
    if fixed.geom_type == "GeometryCollection":
        parts = [g for g in fixed.geoms if not g.is_empty and g.geom_type in ("Polygon", "MultiPolygon")]
        if not parts:
            return None
        fixed = unary_union(parts)
    if fixed.geom_type not in ("Polygon", "MultiPolygon"):
        return None
    # Dissolved unions can be Shapely-valid but still fail MongoDB geo indexing; simplify lightly.
    simplified = fixed.simplify(0.001, preserve_topology=True)
    if simplified.is_empty or simplified.geom_type not in ("Polygon", "MultiPolygon"):
        simplified = fixed.envelope
    return mapping(simplified)


def fetch_core_stack_geometries(state, district, tehsil, db, api_key):
    headers = {"X-API-Key": api_key}
    params = {"state": state, "district": district, "tehsil": tehsil}

    # MWS geometries
    log.info("Fetching MWS geometries from CoRE Stack...")
    r = requests.get(f"{CORE_STACK_BASE}/get_mws_geometries/", headers=headers, params=params, timeout=60)
    if r.status_code != 200:
        log.warning(f"MWS geometry fetch failed: {r.status_code} {r.text[:200]}")
        return False

    mws_geojson = r.json()
    mws_features = mws_geojson.get("features", [])
    log.info(f"  MWS API response type={mws_geojson.get('type')!r}")
    mws_ops = []
    mws_shapes = []
    for feat in mws_features:
        props = feat.get("properties") or {}
        uid = props.get("uid") or props.get("UID")
        geom = feat.get("geometry")
        if uid and geom:
            mws_ops.append(UpdateOne(
                {"uid": uid, "state": state, "district": district, "tehsil": tehsil},
                {"$set": {
                    "uid": uid, "state": state, "district": district, "tehsil": tehsil,
                    "geometry": geom
                }},
                upsert=True
            ))
            try:
                mws_shapes.append(shape(geom))
            except Exception as exc:
                log.warning(f"  MWS uid={uid}: invalid geometry for dissolve ({exc})")
    if mws_ops:
        db.mws_boundaries.bulk_write(mws_ops)
    _log_geometry_ingest_stats(
        "MWS",
        mws_features,
        id_keys=("uid", "UID"),
        stored=len(mws_ops),
        valid_shapes=len(mws_shapes),
    )

    # Village geometries
    log.info("Fetching village geometries from CoRE Stack...")
    r = requests.get(f"{CORE_STACK_BASE}/get_village_geometries/", headers=headers, params=params, timeout=60)
    if r.status_code != 200:
        log.warning(f"Village geometry fetch failed: {r.status_code} {r.text[:200]}")
    else:
        village_geojson = r.json()
        village_features = village_geojson.get("features", [])
        log.info(f"  Village API response type={village_geojson.get('type')!r}")
        village_ops = []
        village_id_keys = (
            "vill_ID", "village_id", "censuscode2011", "census_code", "vill_id",
            "Village_ID", "CensusCode2011", "village_code",
        )
        for feat in village_features:
            props = feat.get("properties") or {}
            vid = None
            for key in village_id_keys:
                val = props.get(key)
                if val is not None and val != "":
                    vid = val
                    break
            geom = feat.get("geometry")
            if vid and geom:
                try:
                    safe_geom = geojson_for_mongo(shape(geom))
                except Exception as exc:
                    log.warning(f"  Village id={vid}: invalid geometry ({exc})")
                    continue
                if not safe_geom:
                    log.warning(f"  Village id={vid}: geometry unusable after repair")
                    continue
                village_ops.append(UpdateOne(
                    {"village_id": vid},
                    {"$set": {
                        "village_id": vid, "state": state, "district": district, "tehsil": tehsil,
                        "geometry": safe_geom
                    }},
                    upsert=True
                ))
        if village_ops:
            db.village_boundaries.bulk_write(village_ops)
        _log_geometry_ingest_stats(
            "Village",
            village_features,
            id_keys=village_id_keys,
            stored=len(village_ops),
        )
        if not village_ops and village_features:
            all_keys: set[str] = set()
            for feat in village_features:
                all_keys.update((feat.get("properties") or {}).keys())
            log.warning(
                f"  Village: no boundaries stored — property keys seen across "
                f"all {len(village_features)} features: {sorted(all_keys)}"
            )

    # Build dissolved tehsil boundary
    if mws_shapes:
        log.info("Building dissolved tehsil boundary...")
        tehsil_poly = unary_union(mws_shapes)
        tehsil_geom = geojson_for_mongo(tehsil_poly)
        if tehsil_geom:
            try:
                db.tehsil_boundaries.replace_one(
                    {"state": state, "district": district, "tehsil": tehsil},
                    {
                        "state": state, "district": district, "tehsil": tehsil,
                        "geometry": tehsil_geom,
                        "mws_count": len(mws_shapes),
                    },
                    upsert=True,
                )
                log.info(f"  Tehsil boundary built and stored (dissolved from {len(mws_shapes)} MWS shapes)")
            except Exception as exc:
                log.warning(f"  Tehsil boundary not stored (geo index): {str(exc)[:240]}")
        else:
            log.warning("  Tehsil boundary skipped: dissolved geometry invalid after repair")
    else:
        log.warning("  Tehsil boundary skipped: no valid MWS shapes to dissolve")

    return True


# ── MongoDB index setup ───────────────────────────────────────────────────────

def ensure_indexes(db):
    db.mws_data.create_index("uid", unique=True)
    db.mws_data.create_index([("tehsils.state", 1), ("tehsils.district", 1), ("tehsils.tehsil", 1)])
    db.mws_data.create_index([("state", 1), ("district", 1), ("tehsil", 1)])
    db.village_data.create_index("village_id", unique=True)
    db.village_data.create_index([("state", 1), ("district", 1), ("tehsil", 1)])
    try:
        db.mws_boundaries.create_index(
            [("uid", 1), ("state", 1), ("district", 1), ("tehsil", 1)],
            unique=True,
        )
        db.mws_boundaries.create_index([("geometry", "2dsphere")])
        db.village_boundaries.create_index([("geometry", "2dsphere")])
        db.tehsil_boundaries.create_index([("geometry", "2dsphere")])
    except Exception as e:
        log.warning(f"Geospatial index (may already exist): {e}")
    log.info("Indexes ensured")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Ingest tehsil Excel into MongoDB")
    parser.add_argument("--excel", required=True, help="Path to tehsil Excel file")
    parser.add_argument("--state", required=True)
    parser.add_argument("--district", required=True)
    parser.add_argument("--tehsil", required=True)
    parser.add_argument("--skip-geometries", action="store_true",
                        help="Skip CoRE Stack geometry fetch")
    parser.add_argument("--force", action="store_true",
                        help="Re-ingest even if already marked complete in manifest")
    args = parser.parse_args()

    api_key = os.getenv("CORE_STACK_API_KEY", "")
    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    ensure_indexes(db)

    manifest_id = f"{args.state}__{args.district}__{args.tehsil}"
    manifest_doc = db.ingest_manifest.find_one({"_id": manifest_id}) or {}

    if manifest_doc.get("status") == "complete" and not args.force:
        log.info(f"Already complete: {manifest_id}. Use --force to re-ingest.")
        return

    # Update manifest to in-progress
    db.ingest_manifest.replace_one(
        {"_id": manifest_id},
        {
            "_id": manifest_id,
            "state": args.state, "district": args.district, "tehsil": args.tehsil,
            "excel_file": os.path.basename(args.excel),
            "status": "in_progress",
            "started_at": datetime.now(timezone.utc).isoformat(),
        },
        upsert=True
    )

    log.info(f"Loading workbook: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)

    # ── Parse MWS-level data ──────────────────────────────────────────────────
    log.info("Parsing MWS sheets...")

    mws_docs = parse_mws(sheet_to_dict(wb, "mws"))
    parse_stream_order(sheet_to_dict(wb, "stream_order", optional=True), mws_docs)
    parse_connectivity(sheet_to_dict(wb, "mws_connectivity", optional=True), mws_docs)
    parse_terrain(sheet_to_dict(wb, "terrain", optional=True), mws_docs)
    parse_terrain_lulc(sheet_to_dict(wb, "terrain_lulc_slope", optional=True), mws_docs, "slope")
    parse_terrain_lulc(sheet_to_dict(wb, "terrain_lulc_plain", optional=True), mws_docs, "plain")
    parse_dem(sheet_to_dict(wb, "dem", optional=True), mws_docs)
    # drainage_density and restoration_vector are not ingested (unused in diagnosis pipeline)
    # parse_drainage_density(sheet_to_dict(wb, "drainage_density", optional=True), mws_docs)
    parse_aquifer(sheet_to_dict(wb, "aquifer_vector", optional=True), mws_docs)
    parse_soge(sheet_to_dict(wb, "soge_vector", optional=True), mws_docs)
    # parse_restoration(sheet_to_dict(wb, "restoration_vector", optional=True), mws_docs)
    parse_river(sheet_to_dict(wb, "river", optional=True), mws_docs)
    parse_canal(sheet_to_dict(wb, "canal", optional=True), mws_docs)

    parse_change_detection(
        sheet_to_dict(wb, "change_detection_degradation"), mws_docs, "degradation",
        {
            "farm_to_barren_ha": "farm_to_barren_area_in_ha",
            "farm_to_built_up_ha": "farm_to_built_up_area_in_ha",
            "farm_to_farm_ha": "farm_to_farm_area_in_ha",
            "farm_to_scrubland_ha": "farm_to_scrub_land_area_in_ha",
            "total_ha": "total_degradation_area_in_ha",
        }
    )
    parse_change_detection(
        sheet_to_dict(wb, "change_detection_deforestation"), mws_docs, "deforestation",
        {
            "forest_to_barren_ha": "forest_to_barren_area_in_ha",
            "forest_to_built_up_ha": "forest_to_built_up_area_in_ha",
            "forest_to_farm_ha": "forest_to_farm_area_in_ha",
            "forest_to_forest_ha": "forest_to_forest_area_in_ha",
            "forest_to_scrubland_ha": "forest_to_scrub_land_area_in_ha",
            "total_ha": "total_deforestation_area_in_ha",
        }
    )
    parse_change_detection(
        sheet_to_dict(wb, "change_detection_urbanization"), mws_docs, "urbanization",
        {
            "barren_shrub_to_built_up_ha": "barren_shrub_to_built_up_area_in_ha",
            "built_up_to_built_up_ha": "built_up_to_built_up_area_in_ha",
            "tree_farm_to_built_up_ha": "tree_farm_to_built_up_area_in_ha",
            "water_to_built_up_ha": "water_to_built_up_area_in_ha",
            "total_ha": "total_urbanization_area_in_ha",
        }
    )
    parse_change_detection(
        sheet_to_dict(wb, "change_detection_afforestation"), mws_docs, "afforestation",
        {
            "barren_to_forest_ha": "barren_to_forest_area_in_ha",
            "built_up_to_forest_ha": "built_up_to_forest_area_in_ha",
            "farm_to_forest_ha": "farm_to_forest_area_in_ha",
            "forest_to_forest_ha": "forest_to_forest_area_in_ha",
            "scrubland_to_forest_ha": "scrub_land_to_forest_area_in_ha",
            "total_ha": "total_afforestation_area_in_ha",
        }
    )
    parse_change_detection(
        sheet_to_dict(wb, "change_detection_cropintensity"), mws_docs, "crop_intensity",
        {
            "single_to_single_ha": "single_to_single_area_in_ha",
            "single_to_double_ha": "single_to_double_area_in_ha",
            "single_to_triple_ha": "single_to_triple_area_in_ha",
            "double_to_single_ha": "double_to_single_area_in_ha",
            "double_to_double_ha": "double_to_double_area_in_ha",
            "double_to_triple_ha": "double_to_triple_area_in_ha",
            "triple_to_single_ha": "triple_to_single_area_in_ha",
            "triple_to_double_ha": "triple_to_double_area_in_ha",
            "triple_to_triple_ha": "triple_to_triple_area_in_ha",
            "total_change_ha": "total_change_crop_intensity_area_in_ha",
        }
    )

    parse_hydrological_annual(sheet_to_dict(wb, "hydrological_annual"), mws_docs)
    parse_hydrological_seasonal(sheet_to_dict(wb, "hydrological_seasonal"), mws_docs)
    parse_swb_annual(sheet_to_dict(wb, "surfaceWaterBodies_annual"), mws_docs)
    parse_lulc_vector(sheet_to_dict(wb, "lulc_vector"), mws_docs)
    parse_cropping_intensity(sheet_to_dict(wb, "croppingIntensity_annual"), mws_docs)
    parse_drought_kharif(sheet_to_dict(wb, "croppingDrought_kharif"), mws_docs)
    parse_drought_causality(sheet_to_dict(wb, "drought_causality"), mws_docs)
    parse_nrega_mws(sheet_to_dict(wb, "nrega_annual", optional=True), mws_docs)
    parse_mws_intersect_villages(sheet_to_dict(wb, "mws_intersect_villages", optional=True), mws_docs)
    parse_mws_intersect_swb(sheet_to_dict(wb, "mws_intersect_swb", optional=True), mws_docs)
    parse_agroecological(sheet_to_dict(wb, "agroecological", optional=True), mws_docs)

    finalize_mws_docs(mws_docs)

    tehsil_ref = {"state": args.state, "district": args.district, "tehsil": args.tehsil}

    # Tag every MWS doc with location metadata (tehsils accumulates across ingests)
    for uid, doc in mws_docs.items():
        doc["state"] = args.state
        doc["district"] = args.district
        doc["tehsil"] = args.tehsil

    # Write MWS docs to MongoDB
    log.info(f"Writing {len(mws_docs)} MWS documents to MongoDB...")
    mws_ops = []
    for uid, doc in mws_docs.items():
        payload = bson_safe(doc)
        mws_ops.append(
            UpdateOne(
                {"uid": uid},
                {
                    "$set": payload,
                    "$addToSet": {"tehsils": tehsil_ref},
                },
                upsert=True,
            )
        )
    if mws_ops:
        result = db.mws_data.bulk_write(mws_ops)
        log.info(f"  MWS: {result.upserted_count} inserted, {result.modified_count} updated")

    # ── Parse village-level data ──────────────────────────────────────────────
    log.info("Parsing village sheets...")
    village_docs = parse_social_economic(wb)
    parse_facilities_proximity(wb, village_docs)
    parse_nrega_village(wb, village_docs)

    # Tag every village doc with tehsil metadata
    for vid, doc in village_docs.items():
        doc.setdefault("state", args.state)
        doc.setdefault("district", args.district)
        doc.setdefault("tehsil", args.tehsil)

    log.info(f"Writing {len(village_docs)} village documents to MongoDB...")
    village_ops = [
        UpdateOne({"village_id": vid}, {"$set": bson_safe(doc)}, upsert=True)
        for vid, doc in village_docs.items()
    ]
    if village_ops:
        result = db.village_data.bulk_write(village_ops)
        log.info(f"  Villages: {result.upserted_count} inserted, {result.modified_count} updated")

    # ── Fetch geometries ──────────────────────────────────────────────────────
    geometries_ok = False
    if not args.skip_geometries:
        if not api_key:
            log.warning("CORE_STACK_API_KEY not set. Skipping geometry fetch.")
        else:
            geometries_ok = fetch_core_stack_geometries(
                args.state, args.district, args.tehsil, db, api_key
            )
    else:
        log.info("Skipping geometry fetch (--skip-geometries)")

    # ── NBSS-LUP AER tagging ──────────────────────────────────────────────────
    if geometries_ok:
        aer_report = validate_aer_geojson()
        if not aer_report.ok:
            log.warning(
                "AER GeoJSON validation failed (%s). "
                "Run scripts/maintenance/fetch_aer_geojson.py before ingest.",
                "; ".join(aer_report.summary_lines()[-2:]),
            )
        else:
            uids = list(mws_docs.keys())
            aer_stats = attach_aer_to_mws(db, uids)
            log.info(
                "AER tagging: updated=%s missing_boundary=%s lookup_failed=%s",
                aer_stats["updated"],
                aer_stats["missing_boundary"],
                aer_stats["lookup_failed"],
            )
            aquifer_stats = refine_aquifer_acwadam(db, uids)
            log.info(
                "Aquifer ACWADAM refine: updated=%s missing_lithology=%s",
                aquifer_stats["updated"],
                aquifer_stats["missing"],
            )

    # ── Update manifest ───────────────────────────────────────────────────────
    db.ingest_manifest.update_one(
        {"_id": manifest_id},
        {"$set": {
            "status": "complete",
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "mws_count": len(mws_docs),
            "village_count": len(village_docs),
            "geometries_fetched": geometries_ok,
        }}
    )
    log.info(f"Done. Manifest updated: {manifest_id}")
    client.close()


if __name__ == "__main__":
    main()
