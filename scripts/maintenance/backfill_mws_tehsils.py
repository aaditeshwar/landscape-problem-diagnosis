#!/usr/bin/env python3
"""Backfill mws_data.tehsils from raw Excel mws sheets (one UID may map to many tehsils)."""

from __future__ import annotations

import argparse
import logging
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from pymongo import MongoClient, UpdateOne

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "runtime"))

from services.tehsil_refs import make_tehsil_ref, merge_tehsil_refs, tehsil_key  # noqa: E402

load_dotenv(ROOT / ".env")

RAW_EXCEL_DIR = ROOT / "data" / "raw_excel"
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017")
DB_NAME = os.getenv("MONGO_DB", "diagnosis_db")

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)


def parse_mws_uids(excel_path: Path) -> list[str]:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if "mws" not in wb.sheetnames:
            return []
        ws = wb["mws"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        if "UID" not in headers:
            return []
        uid_idx = headers.index("UID")
        uids: list[str] = []
        for row in rows[1:]:
            if uid_idx >= len(row):
                continue
            uid = row[uid_idx]
            if uid is None or str(uid).strip() == "":
                continue
            uids.append(str(uid).strip())
        return uids
    finally:
        wb.close()


def tehsil_ref_from_manifest(doc: dict) -> dict | None:
    state = doc.get("state")
    district = doc.get("district")
    tehsil = doc.get("tehsil")
    if state and district and tehsil:
        return make_tehsil_ref(str(state), str(district), str(tehsil))
    parts = str(doc.get("_id", "")).split("__")
    if len(parts) == 3:
        return make_tehsil_ref(parts[0], parts[1], parts[2])
    return None


def iter_tehsil_sources(db, excel_dir: Path):
    seen_manifest: set[str] = set()
    for manifest in db.ingest_manifest.find({"status": "complete"}).sort("_id", 1):
        ref = tehsil_ref_from_manifest(manifest)
        if ref is None:
            continue
        manifest_id = manifest["_id"]
        seen_manifest.add(str(manifest_id))
        excel_path = excel_dir / f"{manifest_id}_data.xlsx"
        if not excel_path.is_file():
            log.warning("Missing excel for manifest %s: %s", manifest_id, excel_path)
            continue
        yield ref, excel_path

    for path in sorted(excel_dir.glob("*__*__*_data.xlsx")):
        manifest_id = path.name[: -len("_data.xlsx")]
        if manifest_id in seen_manifest:
            continue
        parts = manifest_id.split("__")
        if len(parts) != 3:
            continue
        ref = make_tehsil_ref(parts[0], parts[1], parts[2])
        yield ref, path


def build_uid_tehsil_map(db, excel_dir: Path) -> dict[str, list[dict]]:
    uid_refs: dict[str, list[dict]] = defaultdict(list)
    tehsil_files = 0
    for ref, excel_path in iter_tehsil_sources(db, excel_dir):
        tehsil_files += 1
        uids = parse_mws_uids(excel_path)
        log.info("  %s: %s UIDs from %s", tehsil_key(ref), len(uids), excel_path.name)
        for uid in uids:
            uid_refs[uid] = merge_tehsil_refs(uid_refs[uid], ref)
    log.info("Scanned %s tehsil excel files", tehsil_files)
    return uid_refs


def repair_boundary_tehsils(db, *, dry_run: bool = False) -> tuple[int, int]:
    """Ensure mws_boundaries has one row per (uid, tehsil) using existing geometry."""
    boundary_ops: list[UpdateOne] = []
    missing_geometry = 0
    for doc in db.mws_data.find({"tehsils.0": {"$exists": True}}, {"uid": 1, "tehsils": 1}):
        uid = doc.get("uid")
        refs = doc.get("tehsils") or []
        if not uid or not refs:
            continue
        source = db.mws_boundaries.find_one({"uid": uid, "geometry": {"$exists": True}})
        if not source or not source.get("geometry"):
            missing_geometry += 1
            continue
        geom = source["geometry"]
        for ref in refs:
            boundary_ops.append(
                UpdateOne(
                    {
                        "uid": uid,
                        "state": ref["state"],
                        "district": ref["district"],
                        "tehsil": ref["tehsil"],
                    },
                    {
                        "$set": {
                            "uid": uid,
                            "state": ref["state"],
                            "district": ref["district"],
                            "tehsil": ref["tehsil"],
                            "geometry": geom,
                        }
                    },
                    upsert=True,
                )
            )
    if dry_run:
        log.info(
            "Dry run — would upsert %s mws_boundaries rows (%s UIDs missing geometry)",
            len(boundary_ops),
            missing_geometry,
        )
        return len(boundary_ops), missing_geometry
    if boundary_ops:
        result = db.mws_boundaries.bulk_write(boundary_ops, ordered=False)
        log.info(
            "Repaired mws_boundaries: upserted=%s modified=%s (UIDs missing geometry=%s)",
            result.upserted_count,
            result.modified_count,
            missing_geometry,
        )
    else:
        log.info("No boundary repairs needed (UIDs missing geometry=%s)", missing_geometry)
    return len(boundary_ops), missing_geometry


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Report only; do not write MongoDB")
    parser.add_argument(
        "--repair-boundaries",
        action="store_true",
        help="After tehsil backfill, copy each UID geometry to every tehsil in its tehsils list",
    )
    parser.add_argument("--excel-dir", type=Path, default=RAW_EXCEL_DIR)
    args = parser.parse_args()

    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]

    uid_refs = build_uid_tehsil_map(db, args.excel_dir)
    multi = {uid: refs for uid, refs in uid_refs.items() if len(refs) > 1}
    log.info("UIDs seen in excels: %s", len(uid_refs))
    log.info("UIDs in 2+ tehsils: %s", len(multi))
    for uid, refs in sorted(multi.items(), key=lambda item: (-len(item[1]), item[0]))[:15]:
        names = ", ".join(r["tehsil"] for r in refs)
        log.info("  %s -> %s", uid, names)

    mongo_uids = {doc["uid"] for doc in db.mws_data.find({}, {"uid": 1}) if doc.get("uid")}
    excel_uids = set(uid_refs)
    missing_in_mongo = sorted(excel_uids - mongo_uids)
    missing_in_excel = sorted(mongo_uids - excel_uids)
    if missing_in_mongo:
        log.warning("UIDs in excels but not MongoDB: %s (showing 10)", len(missing_in_mongo))
        for uid in missing_in_mongo[:10]:
            log.warning("  %s", uid)
    if missing_in_excel:
        log.warning("UIDs in MongoDB but not in any scanned excel: %s", len(missing_in_excel))

    ops: list[UpdateOne] = []
    for uid, refs in uid_refs.items():
        if uid not in mongo_uids:
            continue
        primary = refs[0]
        ops.append(
            UpdateOne(
                {"uid": uid},
                {
                    "$set": {
                        "tehsils": refs,
                        "state": primary["state"],
                        "district": primary["district"],
                        "tehsil": primary["tehsil"],
                    }
                },
            )
        )

    if args.dry_run:
        log.info("Dry run — would update %s mws_data documents", len(ops))
        if args.repair_boundaries:
            repair_boundary_tehsils(db, dry_run=True)
        client.close()
        return 0

    if not ops:
        log.info("Nothing to update")
        client.close()
        return 0

    result = db.mws_data.bulk_write(ops, ordered=False)
    log.info(
        "Updated tehsils on mws_data: matched=%s modified=%s",
        result.matched_count,
        result.modified_count,
    )
    if args.repair_boundaries:
        repair_boundary_tehsils(db, dry_run=False)
    client.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
