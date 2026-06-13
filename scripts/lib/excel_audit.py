"""Audit tehsil Excel exports against Darwha reference and CoRE Stack geometry APIs."""

from __future__ import annotations

import os
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from shapely.geometry import shape
from shapely.ops import unary_union
from shapely.validation import make_valid

from lib.tehsil_excel_catalog import RAW_EXCEL_DIR, TehsilLocation

ROOT = Path(__file__).resolve().parents[2]
REF = RAW_EXCEL_DIR / "Maharashtra__Yavatmal__Darwha_data.xlsx"
CORE_STACK_BASE = "https://geoserver.core-stack.org/api/v1"

STANDARD_SHEETS = {
    "agroecological",
    "stream_order",
    "mws_connectivity",
    "terrain_lulc_slope",
    "terrain_lulc_plain",
    "river",
    "canal",
}


def sheet_names(path: Path) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    names = set(workbook.sheetnames)
    workbook.close()
    return names


def count_data_rows(path: Path, sheet: str, first_col: int = 0) -> int | None:
    names = sheet_names(path)
    if sheet not in names:
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    count = 0
    for row in workbook[sheet].iter_rows(min_row=2, values_only=True):
        if row and row[first_col]:
            count += 1
    workbook.close()
    return count


def check_mws_union(features: list) -> list[str]:
    issues: list[str] = []
    shapes = []
    invalid = 0
    for feature in features:
        geometry = feature.get("geometry")
        if not geometry:
            continue
        try:
            polygon = shape(geometry)
            if not polygon.is_valid:
                invalid += 1
            shapes.append(polygon)
        except Exception:
            invalid += 1
    if invalid:
        issues.append(f"{invalid} MWS feature(s) have invalid/unparseable geometry")
    if not shapes:
        return issues

    union = unary_union(shapes)
    if not union.is_valid:
        issues.append("Union of MWS polygons is not valid (self-intersections or topology errors)")
    fixed = make_valid(union)
    simplified = fixed.simplify(0.001, preserve_topology=True)
    if not union.is_valid or not union.equals(fixed):
        issues.append(
            "Dissolved tehsil boundary from MWS geometries fails standard geo indexing "
            "(requires make_valid + simplify before storage)"
        )
    if simplified.is_empty:
        issues.append("Dissolved tehsil boundary becomes empty after repair")
    return issues


def check_village_api(features: list, excel_villages: int | None) -> list[str]:
    issues: list[str] = []
    id_keys = (
        "vill_ID",
        "village_id",
        "censuscode2011",
        "census_code",
        "vill_id",
        "Village_ID",
        "CensusCode2011",
        "village_code",
    )
    with_id_geom = 0
    missing_id = 0
    missing_geom = 0
    for feature in features:
        props = feature.get("properties") or {}
        geometry = feature.get("geometry")
        village_id = next((props.get(key) for key in id_keys if props.get(key) is not None), None)
        if not village_id:
            missing_id += 1
        if not geometry:
            missing_geom += 1
        if village_id and geometry:
            with_id_geom += 1
    if missing_id:
        issues.append(f"{missing_id} village feature(s) missing a recognizable village ID field")
    if missing_geom:
        issues.append(f"{missing_geom} village feature(s) missing geometry")
    if excel_villages is not None and with_id_geom < excel_villages:
        issues.append(
            f"Village geometry API returns fewer matchable villages ({with_id_geom}) "
            f"than Excel social_economic_indicator rows ({excel_villages})"
        )
    if excel_villages is not None and len(features) > excel_villages + missing_id:
        issues.append(
            f"Village geometry API returns more features ({len(features)}) "
            f"than Excel village rows ({excel_villages})"
        )
    return issues


def audit_excel_file(
    path: Path,
    location: TehsilLocation,
    *,
    api_key: str | None = None,
    reference_path: Path = REF,
    timeout: int = 90,
) -> dict:
    load_dotenv(ROOT / ".env")
    key = (api_key or os.getenv("CORE_STACK_API_KEY", "")).strip()
    if not key:
        raise RuntimeError("CORE_STACK_API_KEY not set")

    ref_sheets = sheet_names(reference_path) if reference_path.is_file() else set()
    sheets = sheet_names(path)
    missing_standard = sorted(STANDARD_SHEETS - sheets)
    extra_vs_ref = sorted(sheets - ref_sheets) if ref_sheets else []
    missing_vs_ref = sorted(ref_sheets - sheets) if ref_sheets else []

    mws_rows = count_data_rows(path, "mws")
    village_rows = count_data_rows(path, "social_economic_indicator")

    excel_issues: list[str] = []
    if missing_standard:
        excel_issues.append(
            "Missing standard sheet(s) present in Darwha reference export: "
            + ", ".join(missing_standard)
        )
    if missing_vs_ref:
        excel_issues.append(
            "Missing sheet(s) compared to Darwha reference: " + ", ".join(missing_vs_ref)
        )
    if extra_vs_ref:
        excel_issues.append(
            "Extra sheet(s) not in Darwha reference: " + ", ".join(extra_vs_ref)
        )

    api_issues: list[str] = []
    headers = {"X-API-Key": key}
    params = {
        "state": location.state,
        "district": location.district,
        "tehsil": location.tehsil,
    }

    response = requests.get(
        f"{CORE_STACK_BASE}/get_mws_geometries/",
        headers=headers,
        params=params,
        timeout=timeout,
    )
    if response.status_code != 200:
        api_issues.append(f"get_mws_geometries HTTP {response.status_code}: {response.text[:160]}")
    else:
        mws_features = response.json().get("features", [])
        if mws_rows is not None and len(mws_features) != mws_rows:
            api_issues.append(
                f"MWS geometry count mismatch: Excel mws sheet has {mws_rows} rows, "
                f"API returns {len(mws_features)} features"
            )
        api_issues.extend(check_mws_union(mws_features))

    response = requests.get(
        f"{CORE_STACK_BASE}/get_village_geometries/",
        headers=headers,
        params=params,
        timeout=timeout,
    )
    if response.status_code != 200:
        api_issues.append(f"get_village_geometries HTTP {response.status_code}: {response.text[:160]}")
    else:
        village_features = response.json().get("features", [])
        api_issues.extend(check_village_api(village_features, village_rows))

    return {
        "manifest_id": location.manifest_id,
        "state": location.state,
        "district": location.district,
        "tehsil": location.tehsil,
        "excel_file": path.name,
        "excel_path": str(path),
        "excel_mws_rows": mws_rows,
        "excel_village_rows": village_rows,
        "excel_export_issues": excel_issues,
        "core_stack_api_issues": api_issues,
    }


def audit_locations(
    entries: list[tuple[TehsilLocation, Path]],
    *,
    api_key: str | None = None,
) -> dict[str, dict]:
    report: dict[str, dict] = {}
    for location, path in entries:
        report[location.standard_filename] = audit_excel_file(path, location, api_key=api_key)
    return report
